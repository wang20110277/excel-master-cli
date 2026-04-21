"""End-to-end integration tests."""

import json
import pytest
from pathlib import Path
from click.testing import CliRunner
from openpyxl import load_workbook

from excel_master.cli import main
from excel_master.cache import save_cache, load_cache, delete_cache, CACHE_DIR


@pytest.fixture
def runner():
    return CliRunner()


class TestE2ETestCase:
    """Task 8.1: E2E test for test-case template from various inputs."""

    def test_markdown_to_test_case(self, runner, tmp_path):
        md = tmp_path / "cases.md"
        out = tmp_path / "result.xlsx"
        md.write_text(
            "# 测试用例\n\n"
            "## 登录模块\n\n"
            "| 用例编号 | 用例标题 | 优先级 | 前置条件 | 操作步骤 | 预期结果 | 状态 |\n"
            "|---|---|---|---|---|---|---|\n"
            "| TC-001 | 正常登录 | 高 | 用户已注册 | 1.打开登录页\\n2.输入账号密码\\n3.点击登录 | 登录成功 | 待执行 |\n"
            "| TC-002 | 密码错误 | 中 | 用户已注册 | 输入错误密码 | 提示密码错误 | 待执行 |\n"
        )
        result = runner.invoke(main, [
            "create", "--template", "test-case",
            "--input", str(md), "--output", str(out),
        ])
        assert result.exit_code == 0
        assert out.exists()

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
        assert ws.cell(row=2, column=3).value == "正常登录"
        assert ws.cell(row=3, column=3).value == "密码错误"

    def test_json_to_test_case(self, runner, tmp_path):
        data_file = tmp_path / "cases.json"
        out = tmp_path / "result.xlsx"
        data_file.write_text(json.dumps([
            {"title": "Login", "module": "Auth", "priority": "高", "expected": "Success"},
            {"title": "Logout", "module": "Auth", "priority": "低", "expected": "Redirect"},
        ]))
        result = runner.invoke(main, [
            "create", "--template", "test-case",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 3
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=3, column=1).value == "TC-002"

    def test_yaml_to_test_case(self, runner, tmp_path):
        data_file = tmp_path / "cases.yaml"
        out = tmp_path / "result.xlsx"
        data_file.write_text(
            "- title: Search\n  module: Feature\n  expected: Results shown\n"
            "- title: Filter\n  module: Feature\n  expected: Filtered list\n"
        )
        result = runner.invoke(main, [
            "create", "--template", "test-case",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=3).value == "Search"
        assert ws.cell(row=3, column=3).value == "Filter"


class TestE2EOtherTemplates:
    """Task 8.2: E2E test for requirements and project-plan templates."""

    def test_requirements_template(self, runner, tmp_path):
        data_file = tmp_path / "reqs.json"
        out = tmp_path / "reqs.xlsx"
        data_file.write_text(json.dumps([
            {"description": "用户登录", "source": "PRD", "priority": "高", "owner": "Alice"},
            {"description": "密码重置", "source": "PRD", "priority": "中", "owner": "Bob", "status": "开发中"},
        ]))
        result = runner.invoke(main, [
            "create", "--template", "requirements",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "用户登录"
        assert ws.cell(row=2, column=6).value == "Alice"
        assert ws.cell(row=3, column=5).value == "开发中"

    def test_project_plan_template(self, runner, tmp_path):
        data_file = tmp_path / "plan.json"
        out = tmp_path / "plan.xlsx"
        data_file.write_text(json.dumps([
            {"name": "需求分析", "owner": "Alice", "start_date": "2026-04-01", "end_date": "2026-04-15", "status": "已完成"},
            {"name": "开发", "owner": "Bob", "start_date": "2026-04-16", "end_date": "2026-05-15", "status": "进行中"},
        ]))
        result = runner.invoke(main, [
            "create", "--template", "project-plan",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "需求分析"
        assert ws.cell(row=3, column=2).value == "开发"


class TestE2ECustomTemplate:
    """Task 8.3: E2E test for custom template workflow."""

    def test_custom_template_flow(self, runner, tmp_path, monkeypatch):
        monkeypatch.setattr(Path, "home", lambda: tmp_path)

        # Step 1: Init custom template
        result = runner.invoke(main, ["template-init", "simple-report"])
        assert result.exit_code == 0

        # Step 2: Verify scaffold files exist
        tmpl_dir = tmp_path / ".epm" / "templates" / "simple-report"
        assert (tmpl_dir / "template.xlsx").exists()
        assert (tmpl_dir / "schema.yaml").exists()

        # Step 3: Use it
        data_file = tmp_path / "data.json"
        out = tmp_path / "report.xlsx"
        data_file.write_text(json.dumps([{"field1": "Hello"}]))

        result = runner.invoke(main, [
            "create", "--template", "simple-report",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0
        assert out.exists()

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Hello"


class TestE2ECacheResume:
    """Task 8.4: E2E test for cache resume flow."""

    def test_resume_flow(self, runner, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)

        data_file = tmp_path / "data.json"
        out = tmp_path / "output.xlsx"
        records = [{"title": f"Test {i}", "module": "M1", "expected": f"R{i}"} for i in range(5)]

        # First run - partial (simulate cache from previous run)
        data_file.write_text(json.dumps(records))
        save_cache(str(data_file.resolve()), "test-case", records[:3], 3)

        result = runner.invoke(main, [
            "create", "--template", "test-case",
            "--input", str(data_file), "--output", str(out),
        ])
        assert result.exit_code == 0
        assert "Resuming from cache" in result.output

        wb = load_workbook(str(out))
        ws = wb.active
        # 3 cached + 5 new = 8 total data rows
        assert ws.max_row == 9

    def test_clean_flag_ignores_cache(self, runner, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)

        data_file = tmp_path / "data.json"
        out = tmp_path / "output.xlsx"
        records = [{"title": f"Test {i}", "module": "M1", "expected": f"R{i}"} for i in range(3)]

        data_file.write_text(json.dumps(records))
        save_cache(str(data_file.resolve()), "test-case", records[:1], 1)

        result = runner.invoke(main, [
            "create", "--template", "test-case",
            "--input", str(data_file), "--output", str(out), "--clean",
        ])
        assert result.exit_code == 0
        assert "Resuming" not in result.output

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 4  # header + 3 data
