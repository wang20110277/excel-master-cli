"""Tests for the Excel renderer."""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from excel_master.template.registry import TemplateRegistry
from excel_master.renderer.xlsx import XlsxRenderer


@pytest.fixture
def registry():
    return TemplateRegistry()


class TestRenderer:
    def test_basic_render(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [
            {"id": "TC-001", "title": "Login test", "priority": "高",
             "module": "Auth", "expected": "Success"},
        ]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        assert tmp_output.exists()
        wb = load_workbook(str(tmp_output))
        ws = wb.active
        # Header row + 1 data row
        assert ws.max_row == 2
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=2, column=3).value == "Login test"

    def test_auto_generate_ids(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [
            {"title": "Test A", "module": "M1", "expected": "R1"},
            {"title": "Test B", "module": "M1", "expected": "R2"},
        ]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "TC-001"
        assert ws.cell(row=3, column=1).value == "TC-002"

    def test_default_values(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [{"title": "Test", "module": "M1", "expected": "R1"}]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        # priority (col 4) default = 中, status (col 8) default = 待执行
        assert ws.cell(row=2, column=4).value == "中"
        assert ws.cell(row=2, column=8).value == "待执行"

    def test_validate_invalid_value(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [
            {"title": "Test", "module": "M1", "priority": "紧急", "expected": "R1"},
        ]
        renderer = XlsxRenderer(tmpl)
        with pytest.warns(UserWarning, match="Invalid value"):
            renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        # Invalid priority replaced with default "中"
        assert ws.cell(row=2, column=4).value == "中"

    def test_multiline_row_height(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [
            {"title": "Test", "module": "M1", "steps": "1. Step1\n2. Step2\n3. Step3", "expected": "R1"},
        ]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.row_dimensions[2].height >= 54  # 3 lines * 18

    def test_style_inheritance(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [{"title": "Test", "module": "M1", "expected": "R1"}]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        data_cell = ws.cell(row=2, column=1)
        # Data cells should NOT be bold
        assert data_cell.font.bold is False
        # Data cells should have border
        assert data_cell.border.left.style == "thin"

    def test_multiple_records(self, registry, tmp_output):
        tmpl = registry.get_template("test-case")
        records = [
            {"id": f"TC-{i:03d}", "title": f"Test {i}", "module": "M1", "expected": f"R{i}"}
            for i in range(1, 6)
        ]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.max_row == 6  # 1 header + 5 data

    def test_requirements_template(self, registry, tmp_output):
        tmpl = registry.get_template("requirements")
        records = [{"description": "User login", "owner": "Alice"}]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "User login"
        assert ws.cell(row=2, column=6).value == "Alice"

    def test_project_plan_template(self, registry, tmp_output):
        tmpl = registry.get_template("project-plan")
        records = [{"name": "Design DB", "owner": "Bob", "status": "进行中"}]
        renderer = XlsxRenderer(tmpl)
        renderer.render(records, tmp_output)

        wb = load_workbook(str(tmp_output))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "Design DB"
        assert ws.cell(row=2, column=6).value == "进行中"
