"""Excel Master CLI - command definitions."""

from __future__ import annotations

import sys
from pathlib import Path

import click

from excel_master import __version__
from excel_master.cache import clean_stale_caches, delete_cache, load_cache
from excel_master.parser import get_parser, auto_detect_format
from excel_master.renderer.xlsx import XlsxRenderer
from excel_master.template.registry import TemplateRegistry


@click.group()
@click.version_option(version=__version__, prog_name="epm")
def main():
    """Excel Master CLI - Generate structured Excel documents from text inputs."""


@main.command("list-templates")
def list_templates():
    """List all available templates (built-in + user-defined)."""
    registry = TemplateRegistry()
    templates = registry.list_templates()
    if not templates:
        click.echo("No templates found.")
        return

    click.echo(f"{'Name':<20} {'Display Name':<15} {'Source':<10} Description")
    click.echo("-" * 70)
    for t in templates:
        click.echo(f"{t['name']:<20} {t['display_name']:<15} {t['source']:<10} {t['description']}")


@main.command("show-schema")
@click.argument("template_name")
def show_schema(template_name):
    """Show the input data structure for a template."""
    registry = TemplateRegistry()
    try:
        schema = registry.get_schema(template_name)
    except ValueError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)

    click.echo(f"Template: {schema.display_name} ({schema.name})")
    click.echo(f"Description: {schema.description}")
    click.echo()
    click.echo(f"{'Field':<18} {'Header':<12} {'Width':<6} {'Required':<9} {'Default':<10} Notes")
    click.echo("-" * 80)

    for col in schema.columns:
        notes = []
        if col.auto_generate:
            notes.append("auto_generate")
        if col.validate:
            notes.append(f"validate: {col.validate}")
        if col.multiline:
            notes.append("multiline")
        if col.extract:
            kws = [kw for e in col.extract for kw in e.get("keywords", [])]
            notes.append(f"keywords: {kws}")
        click.echo(
            f"{col.field:<18} {col.header:<12} {col.width:<6} "
            f"{'Yes' if col.required else 'No':<9} "
            f"{str(col.default or ''):<10} {'; '.join(notes)}"
        )


@main.command("create")
@click.option("--template", "-t", required=True, help="Template name")
@click.option("--input", "-i", "input_path", required=True, help="Input file path or '-' for stdin")
@click.option("--output", "-o", "output_path", default="./output.xlsx", help="Output file path")
@click.option("--format", "-f", "fmt", default=None, help="Input format (md|txt|docx|json|yaml)")
@click.option("--clean", is_flag=True, help="Ignore cache, start from scratch")
def create(template, input_path, output_path, fmt, clean):
    """Generate an Excel file from an input document."""
    # Clean stale caches on startup
    clean_stale_caches()

    # Resolve format
    if input_path == "-":
        if not fmt:
            click.echo("Error: --format is required when reading from stdin", err=True)
            sys.exit(1)
        content = sys.stdin.read()
        source = "<stdin>"
    else:
        source = str(Path(input_path).resolve())
        if not fmt:
            fmt = auto_detect_format(input_path)
        content = None if fmt in ("docx", "image") else Path(input_path).read_text(encoding="utf-8")

    # Load template
    registry = TemplateRegistry()
    try:
        tmpl = registry.get_template(template)
    except ValueError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)

    # Check cache for resume
    existing_records = []
    if not clean and input_path != "-":
        cached = load_cache(source, template)
        if cached and cached.get("records"):
            existing_records = cached["records"]
            click.echo(f"Resuming from cache: {len(existing_records)} records loaded")

    # Parse input
    parser = get_parser(fmt)
    if input_path != "-" and (input_path.endswith(".docx") or fmt == "image"):
        result = parser.parse_file(input_path)
    else:
        result = parser.parse(content, source=source)

    all_records = existing_records + result.records
    click.echo(f"Parsed {len(result.records)} new records (total: {len(all_records)})")

    # Render
    renderer = XlsxRenderer(tmpl)
    output = renderer.render(all_records, output_path)

    # Print any warnings
    for w in renderer.warnings:
        click.echo(f"Warning: {w}", err=True)

    # Delete cache on success
    if input_path != "-":
        delete_cache(source, template)

    click.echo(f"Generated: {output}")


@main.command("template-init")
@click.argument("name")
def template_init(name):
    """Initialize a custom template directory at ~/.epm/templates/<name>/."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    dest = Path.home() / ".epm" / "templates" / name
    if dest.exists():
        click.echo(f"Error: Template '{name}' already exists at {dest}", err=True)
        sys.exit(1)

    dest.mkdir(parents=True, exist_ok=True)

    # Create scaffold xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = name
    header = ws.cell(row=1, column=1, value="字段1")
    header.font = Font(bold=True)
    header.alignment = Alignment(horizontal="center")
    header.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ws.freeze_panes = "A2"
    wb.save(str(dest / "template.xlsx"))

    # Create scaffold schema
    schema_content = f"""name: {name}
display_name: {name}
description: "Custom template"
columns:
  - field: field1
    header: 字段1
    width: 20
"""
    (dest / "schema.yaml").write_text(schema_content, encoding="utf-8")

    click.echo(f"Created template '{name}' at {dest}")
    click.echo(f"  Edit {dest}/template.xlsx for styles")
    click.echo(f"  Edit {dest}/schema.yaml for field mapping")


if __name__ == "__main__":
    main()
