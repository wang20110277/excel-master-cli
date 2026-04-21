"""Template loader - loads .xlsx template + .yaml schema."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


@dataclass
class TemplateColumn:
    field: str
    header: str
    width: int = 15
    required: bool = False
    default: Any = None
    validate: list[str] | None = None
    auto_generate: bool = False
    multiline: bool = False
    extract: list[dict[str, list[str]]] | None = None


@dataclass
class TemplateSchema:
    name: str
    display_name: str
    description: str
    columns: list[TemplateColumn] = field(default_factory=list)

    @property
    def field_map(self) -> dict[str, TemplateColumn]:
        return {c.field: c for c in self.columns}


@dataclass
class Template:
    name: str
    schema: TemplateSchema
    xlsx_path: Path
    workbook: Any = None

    def load_workbook(self):
        if self.workbook is None:
            self.workbook = load_workbook(str(self.xlsx_path))
        return self.workbook


class TemplateLoader:
    """Load a template pair (xlsx + yaml schema)."""

    @staticmethod
    def load_schema(schema_path: Path) -> TemplateSchema:
        with open(schema_path, encoding="utf-8") as f:
            data = yaml.safe_load(f)

        columns = []
        for col_data in data.get("columns", []):
            columns.append(TemplateColumn(
                field=col_data["field"],
                header=col_data["header"],
                width=col_data.get("width", 15),
                required=col_data.get("required", False),
                default=col_data.get("default"),
                validate=col_data.get("validate"),
                auto_generate=col_data.get("auto_generate", False),
                multiline=col_data.get("multiline", False),
                extract=col_data.get("extract"),
            ))

        return TemplateSchema(
            name=data["name"],
            display_name=data["display_name"],
            description=data.get("description", ""),
            columns=columns,
        )

    @staticmethod
    def load(xlsx_path: Path, schema_path: Path) -> Template:
        schema = TemplateLoader.load_schema(schema_path)
        return Template(
            name=schema.name,
            schema=schema,
            xlsx_path=xlsx_path,
        )
