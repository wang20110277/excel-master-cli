"""Excel rendering engine - fills template with parsed records."""

from __future__ import annotations

import copy
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from excel_master.template.loader import Template, TemplateColumn


class XlsxRenderer:
    """Render parsed records into an Excel file using a template."""

    def __init__(self, template: Template):
        self.template = template
        self.schema = template.schema
        self.warnings: list[str] = []

    def render(self, records: list[dict[str, Any]], output_path: str | Path) -> Path:
        """Render records to an Excel file."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(str(self.template.xlsx_path))
        ws = wb.active

        # Get header styles from row 1
        header_styles = self._extract_header_styles(ws)

        auto_counters: dict[str, int] = {}
        for col in self.schema.columns:
            if col.auto_generate:
                auto_counters[col.field] = 0

        for row_idx, record in enumerate(records, start=2):
            for col_idx, col_def in enumerate(self.schema.columns, start=1):
                value = self._resolve_value(record, col_def, auto_counters)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply inherited styles
                if col_idx in header_styles:
                    style = header_styles[col_idx]
                    cell.font = style["font"]
                    cell.border = style["border"]
                    cell.alignment = style["alignment"]
                    # Auto-adjust row height for multiline
                if col_def.multiline and value and isinstance(value, str) and "\n" in value:
                    line_count = value.count("\n") + 1
                    ws.row_dimensions[row_idx].height = max(
                        ws.row_dimensions[row_idx].height or 20,
                        line_count * 18,
                    )

        # Add dropdown validation for columns with validate options
        for col_idx, col_def in enumerate(self.schema.columns, start=1):
            if col_def.validate:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(col_def.validate)}"',
                    allow_blank=True,
                )
                dv.error = f"请从下拉列表中选择有效值: {', '.join(col_def.validate)}"
                dv.errorTitle = "输入无效"
                dv.prompt = f"请选择: {', '.join(col_def.validate)}"
                dv.promptTitle = col_def.header
                last_row = len(records) + 1
                dv.add(f"{col_letter}2:{col_letter}{last_row}")
                ws.add_data_validation(dv)

        wb.save(str(output_path))
        return output_path

    def _resolve_value(
        self,
        record: dict[str, Any],
        col: TemplateColumn,
        auto_counters: dict[str, int],
    ) -> Any:
        """Resolve a cell value from a record for a given column."""
        # Try direct field match
        value = record.get(col.field)

        # Try header name match
        if value is None:
            value = record.get(col.header)

        # Try extract keyword match
        if value is None and col.extract:
            for entry in col.extract:
                for kw in entry.get("keywords", []):
                    if kw in record:
                        value = record[kw]
                        break
                if value is not None:
                    break

        # Auto-generate ID
        if value is None and col.auto_generate:
            auto_counters[col.field] = auto_counters.get(col.field, 0) + 1
            prefix = self._derive_prefix()
            value = f"{prefix}-{auto_counters[col.field]:03d}"

        # Apply default
        if value is None:
            value = col.default

        # Validate
        if value is not None and col.validate and str(value) not in col.validate:
            msg = (
                f"Invalid value '{value}' for field '{col.field}'. "
                f"Allowed: {col.validate}. Using default: {col.default}"
            )
            self.warnings.append(msg)
            warnings.warn(msg, stacklevel=2)
            value = col.default

        # Required field check
        if value is None and col.required:
            msg = f"Missing required field '{col.field}' in record"
            self.warnings.append(msg)
            warnings.warn(msg, stacklevel=2)

        return value

    def _derive_prefix(self) -> str:
        """Derive auto-generate prefix from template name."""
        name = self.schema.name
        parts = name.split("-")
        return "".join(p[0].upper() for p in parts)

    def _extract_header_styles(self, ws) -> dict[int, dict[str, Any]]:
        """Extract styles from the header row (row 1)."""
        styles = {}
        for col_idx in range(1, len(self.schema.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            # Clone font but set bold=False for data rows
            font = copy.copy(cell.font)
            font = Font(
                name=font.name,
                size=font.size,
                bold=False,
                italic=font.italic,
            )
            styles[col_idx] = {
                "font": font,
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
            }
        return styles
